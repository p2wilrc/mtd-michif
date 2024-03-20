#!/usr/bin/env python

"""Read information from TM metadata XLSX and write to JSON for future use."""

import argparse
import json
import logging
import os
from collections import defaultdict
from datetime import datetime
from os import PathLike
from pathlib import Path

from openpyxl import load_workbook  # type: ignore

LOGGER = logging.getLogger(Path(__file__).stem)


def make_argparse():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("metadata", help="Excel notebook with metadata")
    return parser


def read_metadata(path: PathLike) -> dict[str, dict | list]:
    wb = load_workbook(path)

    people = {}
    hdr = list(next(wb["People"].iter_rows(1, 1, values_only=True)))
    for row in wb["People"].iter_rows(2, values_only=True):
        people_data = {hdr[i]: x for i, x in enumerate(row)}
        people[people_data["ID"]] = people_data
    sessions: defaultdict[str, dict] = defaultdict(dict)
    hdr = list(next(wb["Sessions"].iter_rows(1, 1, values_only=True)))
    for row_idx in range(2, wb["Sessions"].max_row + 1):
        row_data = {hdr[i]: x.value for i, x in enumerate(wb["Sessions"][row_idx])}
        if row_data["FileName"] is None:
            continue
        for col in ("SessionDate", "StatusDate"):
            if row_data[col] is not None:
                if isinstance(row_data[col], datetime):
                    row_data[col] = row_data[col].strftime("%Y-%m-%d")  # type: ignore
                else:
                    LOGGER.info("%s is not datetime: %s", col, row_data[col])
        filename, _ = os.path.splitext(str(row_data["FileName"]))
        if filename in sessions:
            LOGGER.warning("Duplicate file %s", filename)
        sessions[filename] = row_data
    notes = []
    hdr = list(
        next(wb["Annotator Notes & Missed Words"].iter_rows(1, 1, values_only=True))
    )
    for row in wb["Annotator Notes & Missed Words"].iter_rows(2, values_only=True):
        row_data = {hdr[i]: x for i, x in enumerate(row)}
        if row_data["FileName"] is None:
            continue
        for col in ("StatusDate",):
            if row_data[col] is not None:
                if isinstance(row_data[col], datetime):
                    row_data[col] = row_data[col].strftime("%Y-%m-%d")  # type: ignore
                else:
                    LOGGER.info("%s is not datetime: %s", col, row_data[col])
        notes.append(row_data)
    return {"people": people, "sessions": sessions, "notes": notes}


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    parser = make_argparse()
    args = parser.parse_args()
    print(json.dumps(read_metadata(args.metadata), indent=2))
